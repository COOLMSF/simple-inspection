import os
import time
import requests
import datetime

import openpyxl
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter


def DBA_CHECKOUT(wb):
    sheet = wb.create_sheet(index=0, title='数据库审计ls巡检')

    orange_fill = PatternFill(fill_type='solid', fgColor="FFC125")
    green_fill = PatternFill(fill_type='solid', fgColor="AACF91")

    info_title_list = ["节点名称", "ip", "授权码", "模板型号",
                       "激活日期", "授权日期", "失效时间", "型号归属"]

    column = 1
    for title in info_title_list:
        sheet.cell(row=1, column=column).value = title
        sheet.cell(row=1, column=column).fill = orange_fill
        column += 1

    device_status_dic = {
        "0": "未分配",
        "1": "分配中",
        "2": "已分配",
        "3": "过期",
        "-3": "过期回收",
        "-1": "异常",
        "4": "扩容中",
    }

    is_expand_dic = {
        "0": "否",
        "1": "是",
    }

    is_mulit = {
        "0": "否",
        "1": "是",
    }

    node_dic = {
        '北京2': '100.64.25.54',
        '北京3': '100.64.41.9',
        # '广州骨干': '100.75.1.24',
        '杭州': '100.66.25.54',
        '石家庄': '100.69.17.54',
        '南昌new': '100.70.17.54',
        '连云港牧野': '100.66.9.54',
        '赛迪紫鸾': '100.70.9.25',
        '唐山': '100.69.69.54',
        # '上海': '100.66.1.24',

    }

    header = {
        "Content-Type": "application/x-www-form-urlencoded",
        "X-Auth-Token": "111111",
    }

    body = {
        'user_name': 'admin',
        'user_pwd': 'Xfhbcx10!',
    }

    row = 2

    for name, ip in node_dic.items():

        sheet.cell(row=row, column=1).value = name
        sheet.cell(row=row, column=1).fill = green_fill

        sheet.cell(row=row, column=2).value = ip
        sheet.cell(row=row, column=2).fill = green_fill

        url = f'https://{ip}/cas/authcodeList'

        DBA_requests = requests.post(url=url, headers=header, verify=False, data=body)

        content = DBA_requests.json()

        data_list = content["data"]

        for data_dic in data_list:
            column = 3
            for key, value in data_dic.items():
                if key in ["authcode", "model", "model_type"]:
                    sheet.cell(row=row, column=column).value = value

                if key in ["create_date", "uptime", "active_date", "grant_date"]:
                    value = int(value)
                    ten_timeArray = time.localtime(value)
                    ten_otherStyleTime = time.strftime("%Y-%m-%d %H:%M:%S", ten_timeArray)
                    sheet.cell(row=row, column=column).value = ten_otherStyleTime

                column += 1
            row += 1
        row += 1

    max_rows = sheet.max_row
    max_col = sheet.max_column

    align = Alignment(horizontal='center', vertical='center')

    for i in range(1, max_rows + 1):
        for j in range(1, max_col + 1):
            sheet.cell(i, j).alignment = align

    # 设置一个字典用于保存列宽数据
    dims = {}

    # 遍历表格数据，获取自适应列宽数据
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                # 遍历整个表格，把该列所有的单元格文本进行长度对比，找出最长的单元格
                # 在对比单元格文本时需要将中文字符识别为1.7个长度，英文字符识别为1个，这里只需要将文本长度直接加上中文字符数量即可
                # re.findall('([\u4e00-\u9fa5])', cell.value)能够识别大部分中文字符
                cell_len = len(str(cell.value)) * 1.5
                dims[cell.column] = max((dims.get(cell.column, 0), cell_len))

    for col, value in dims.items():
        # 设置列宽，get_column_letter用于获取数字列号对应的字母列号，最后值+2是用来调整最终效果的
        sheet.column_dimensions[get_column_letter(col)].width = value + 2


if __name__ == '__main__':

    time_ = datetime.datetime.now().strftime("%Y-%m-%d_%H:%M")

    if not os.path.exists('/data_device'):
        os.mkdir('/data_device')

    if os.path.exists(f'/data_device/{time_}/数据库审计ls'):
        print(f"请删除/data_device/{time_}/数据库审计ls/")
        raise IndexError('存在缓存数据！')

    os.mkdir(f'/data_device/{time_}/')
    os.mkdir(f'/data_device/{time_}/数据库审计ls/')

    data_path = f'/data_device/{time_}/数据库审计ls'

    WB = openpyxl.Workbook()

    # 数据库审计巡检
    DBA_CHECKOUT(WB)

    WB.save(f'{data_path}/数据库审计ls巡检信息.xlsx')

    print(f'输出的文件在：{data_path}')
