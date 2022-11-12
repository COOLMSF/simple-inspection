import requests
import openpyxl
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter


def LOG_CHECKOUT(wb):
    # 修改资产信息
    node_info = {
        "北京2": "100.64.25.58",
        "北京3": "100.64.41.58",
        "广州联通牧野": "100.70.1.58",
        "广州骨干": "100.75.1.58",
        "杭州": "100.66.25.58",
        "河北广电": "100.69.17.58",
        "江西联通": "100.70.17.58",
        "连云港牧野": "100.66.9.58",
        "重庆牧野": "100.72.1.58",
        "赛迪紫鸾": "100.70.9.29",
        "唐山": "100.69.69.58",
        "山西": "100.69.41.58",
        "武汉new": "100.68.35.58",
        "上海": "100.66.1.58",
        "宁夏": "100.68.65.51",
        "四川": "100.68.25.26",
        "天津": "100.64.7.228",
    }

    orange_fill = PatternFill(fill_type='solid', fgColor="FFC125")
    green_fill = PatternFill(fill_type='solid', fgColor="AACF91")
    # 创建一个sheet并返回这个sheet
    sheet = wb.create_sheet(index=0, title='日志审计LS巡检')

    title_info = ["节点", "IP", "授权名称", "授权总数", "已使用授权", "剩余授权"]

    column = 1
    for title in title_info:
        sheet.cell(row=1, column=column).value = title
        sheet.cell(row=1, column=column).fill = orange_fill
        column += 1

    row = 2

    for key, value in node_info.items():
        url = f"https://{value}:28443/licsmgr/api/v1/oauth2/token?"
        body = {
            "username": "admin",
            "password": "WENC(3qm5J8iKqu0Y3kEo/P90kA==)",
            "grant_type": "password",
        }
        header = {
            "Authorization": "Basic Y2xpZW50OmNsaWVudHB3ZA==",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36",
            "Content-Type": "application/x-www-form-urlencoded",
        }

        response = requests.post(url=url, headers=header, data=body, verify=False).json()
        access_token = response.get("access_token")
        token_type = response.get("token_type")

        lic_url = f"https://{value}:28443/licsmgr/api/v1/licenses/statistic?page=1&pageSize=8&pdtName=&licAlias="
        header2 = {
            "User-Agent": "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36",
            "Authorization": f"{token_type} {access_token}",
        }
        response2 = requests.get(url=lic_url, headers=header2, verify=False).json()
        sheet.cell(row=row, column=1).value = key
        sheet.cell(row=row, column=2).value = value
        try:
            info_dic = response2.get("rows")[0]
            licName = info_dic.get("licName")
            licAmount = info_dic.get("licAmount")
            licUsedAmount = info_dic.get("licUsedAmount")
            licRemainAmount = info_dic.get("licRemainAmount")

            sheet.cell(row=row, column=3).value = licName
            sheet.cell(row=row, column=4).value = licAmount
            sheet.cell(row=row, column=5).value = licUsedAmount
            sheet.cell(row=row, column=6).value = licRemainAmount
        except:
            print(f'{key}节点没有授权')

        row += 1

    # 遍历全部表格并设置居中
    max_rows = sheet.max_row
    max_col = sheet.max_column

    # 设置居中属性
    align = Alignment(horizontal='center', vertical='center')

    for i in range(1, max_rows + 1):
        for j in range(1, max_col + 1):
            sheet.cell(i, j).alignment = align

    # 设置宽度自适应
    # 设置一个字典用于保存列宽数据
    dims = {}

    # 遍历表格数据，获取自适应列宽数据
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                # 遍历整个表格，把该列所有的单元格文本进行长度对比，找出最长的单元格
                # 在对比单元格文本时需要将中文字符识别为1.7个长度，英文字符识别为1个，这里只需要将文本长度直接加上中文字符数量即可
                # re.findall('([\u4e00-\u9fa5])', cell.value)能够识别大部分中文字符
                cell_len = len(str(cell.value)) * 1.5
                dims[cell.column] = max((dims.get(cell.column, 0), cell_len))

    for col, value in dims.items():
        # 设置列宽，get_column_letter用于获取数字列号对应的字母列号，最后值+2是用来调整最终效果的
        sheet.column_dimensions[get_column_letter(col)].width = value + 2


if __name__ == '__main__':
    wb = openpyxl.Workbook()
    LOG_CHECKOUT(wb)
    wb.save("日志审计ls.xlsx")
