import os
import requests
import datetime

import openpyxl
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter


def CBH_CHECK(wb):
    # 创建sheet表
    sheet = wb.create_sheet(index=0, title='堡垒机LS巡检')

    orange_fill = PatternFill(fill_type='solid', fgColor="FFC125")

    # 设置标题
    title_list = ['节点名称', 'IP', 'uuid', '授权类型', 'validDays', 'validFrom', 'validTo', 'dev', 'used']
    column = 1
    for title in title_list:
        sheet.cell(row=1, column=column).value = title
        sheet.cell(row=1, column=column).fill = orange_fill
        column += 1

    # 配置节点信息
    node_dic = {
        '北京2': '100.64.25.57',
        '北京3': '100.64.41.57',
        '广州骨干': '100.75.1.57',
        '广州合营': '100.70.1.57',
        '杭州': '100.66.25.57',
        '石家庄': '100.69.17.57',
        '南昌new': '100.70.17.57',
        '连云港牧野': '100.66.9.57',
        '重庆new': '100.72.1.57',
        '赛迪紫鸾': '100.70.8.55',
        '唐山': '100.69.69.57',
        '武汉': '100.68.35.57',
        '上海': '100.66.1.57',
        '宁夏': '100.68.65.50',
        '四川': '100.68.25.25',
    }

    row = 2
    for nodename, ip in node_dic.items():
        CBH_request = requests.get(url=f'https://{ip}/lic/list', verify=False)
        resp_list = CBH_request.json()

        green_fill = PatternFill(fill_type='solid', fgColor="AACF91")

        sheet.cell(row=row, column=1).value = nodename
        sheet.cell(row=row, column=1).fill = green_fill

        sheet.cell(row=row, column=2).value = ip
        sheet.cell(row=row, column=2).fill = green_fill

        for info_dic in resp_list:

            column = 3
            for key, value in info_dic.items():
                sheet.cell(row=row, column=column).value = value
                column += 1

            row += 1
        row += 1

    max_rows = sheet.max_row
    max_col = sheet.max_column

    align = Alignment(horizontal='center', vertical='center')

    for i in range(1, max_rows + 1):
        for j in range(1, max_col + 1):
            sheet.cell(i, j).alignment = align

    # 设置一个字典用于保存列宽数据
    dims = {}

    # 遍历表格数据，获取自适应列宽数据
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                # 遍历整个表格，把该列所有的单元格文本进行长度对比，找出最长的单元格
                # 在对比单元格文本时需要将中文字符识别为1.7个长度，英文字符识别为1个，这里只需要将文本长度直接加上中文字符数量即可
                # re.findall('([\u4e00-\u9fa5])', cell.value)能够识别大部分中文字符
                cell_len = len(str(cell.value)) * 1.4
                dims[cell.column] = max((dims.get(cell.column, 0), cell_len))
    for col, value in dims.items():
        # 设置列宽，get_column_letter用于获取数字列号对应的字母列号，最后值+2是用来调整最终效果的
        sheet.column_dimensions[get_column_letter(col)].width = value + 2


if __name__ == '__main__':

    time = datetime.datetime.now().strftime("%Y-%m-%d_%H:%M")

    if not os.path.exists('/data_device'):
        os.mkdir('/data_device')

    if os.path.exists(f'/data_device/{time}'):
        print(f"请删除/data_device/{time}文件夹")
        raise IndexError('存在缓存数据！')

    os.mkdir(f'/data_device/{time}')

    data_path = f'/data_device/{time}'

    WB = openpyxl.Workbook()

    # 堡垒机巡检
    CBH_CHECK(WB)

    WB.save(f'{data_path}/堡垒机ls巡检信息.xlsx')

    print(f'输出的文件在：{data_path}')
