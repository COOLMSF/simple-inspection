# https://100.69.17.56/api/getToken/
import os
import time
import datetime
import requests
import hashlib

import openpyxl
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter


def md5_hash(data):
    data = data.encode('utf-8')
    return hashlib.md5(data).hexdigest()


def WAF_CHECKOUT(wb):
    sheet = wb.create_sheet(index=0, title='waf_ls')

    title_list = ['节点名称', "ip", "授权型号", "授权类型", "时间", "数量", "对应的授权id", "所属授权包导入时间", "类型"]

    orange_fill = PatternFill(fill_type='solid', fgColor="FFC125")
    green_fill = PatternFill(fill_type='solid', fgColor="AACF91")

    align = Alignment(horizontal='center', vertical='center')

    node_dic = {
    #    '北京3': '100.69.17.56',
        '河北广电': '100.69.17.56',
    #    '赛迪紫鸾': '100.70.9.23',
    #    '宁夏': '100.68.65.55',
    }

    type_dic = {
        "1": "功能",
        "2": "特征库时间",
        "3": "防篡改授权包",
    }

    column = 1
    for title in title_list:
        sheet.cell(row=1, column=column).value = title
        sheet.cell(row=1, column=column).fill = orange_fill
        column += 1

    row = 2
    for name, ip in node_dic.items():

        sheet.cell(row=row, column=1).value = name
        sheet.cell(row=row, column=1).fill = green_fill

        sheet.cell(row=row, column=2).value = ip
        sheet.cell(row=row, column=2).fill = green_fill

        url = f'https://{ip}/api/getToken/'

        url_sub = f'/license/total/time/list/'

        # 拿token
        header = {
            'Content-Type': 'application/x-www-form-urlencoded',
        }

        body = {
            'user_name': 'admin',
            'user_pwd': 'Xfhbcx10!',
        }

        token = requests.post(url=url, data=body, headers=header, verify=False).json()["data"]["token"]

        # 拼接
        content_sgin = token + url_sub

        # 加密
        sgin = md5_hash(content_sgin)

        spec_list = ['W2000-V100-G2', 'W2000-V200-G2', 'W2000-V300-G2']

        type_auth_list = ['formal', 'temporary']

        for spec in spec_list:
            for type_auth in type_auth_list:
                api_url = 'https://%s/api/access/?sign=%s&user_name=admin&url=%s&param={"product":"Waf","spec":"%s","authtype":"%s"}' % (
                    ip, sgin, url_sub, spec, type_auth)
                sheet.cell(row=row, column=3).value = spec
                sheet.cell(row=row, column=4).value = type_auth

                sheet.cell(row=row, column=3).fill = green_fill
                sheet.cell(row=row, column=4).fill = green_fill

                content_dir = requests.post(url=api_url, headers=header, data=body, verify=False).json()
                data_list = content_dir['data']
                if not data_list:
                    row += 1
                else:
                    for data in data_list:

                        column = 5

                        for key, value in data.items():
                            if key == "assignedtime":
                                value = int(value)
                                ten_timeArray = time.localtime(value)
                                ten_otherStyleTime = time.strftime("%Y-%m-%d %H:%M:%S", ten_timeArray)
                                sheet.cell(row=row, column=column).value = ten_otherStyleTime
                            elif key == "typeid":
                                type_auth = type_dic[str(value)]
                                sheet.cell(row=row, column=column).value = type_auth
                            else:
                                sheet.cell(row=row, column=column).value = value

                            column += 1

                        row += 1
        row += 1

    max_rows = sheet.max_row
    max_col = sheet.max_column

    for i in range(1, max_rows + 1):
        for j in range(1, max_col + 1):
            sheet.cell(i, j).alignment = align

    # 设置一个字典用于保存列宽数据
    dims = {}

    # 遍历表格数据，获取自适应列宽数据
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                # 遍历整个表格，把该列所有的单元格文本进行长度对比，找出最长的单元格
                # 在对比单元格文本时需要将中文字符识别为1.7个长度，英文字符识别为1个，这里只需要将文本长度直接加上中文字符数量即可
                # re.findall('([\u4e00-\u9fa5])', cell.value)能够识别大部分中文字符
                cell_len = len(str(cell.value))
                dims[cell.column] = max((dims.get(cell.column, 0), cell_len))

    for col, value in dims.items():
        # 设置列宽，get_column_letter用于获取数字列号对应的字母列号，最后值+2是用来调整最终效果的
        sheet.column_dimensions[get_column_letter(col)].width = value + 10


if __name__ == '__main__':

    time_ = datetime.datetime.now().strftime("%Y-%m-%d_%H:%M")

    if not os.path.exists('/data_device'):
        os.mkdir('/data_device')

    if os.path.exists(f'/data_device/{time_}/WAF_LS'):
        print(f"请删除/data_device/{time_}/WAF_LS/")
        raise IndexError('存在缓存数据！')

    os.mkdir(f'/data_device/{time_}/')
    os.mkdir(f'/data_device/{time_}/WAF_LS')

    data_path = f'/data_device/{time_}/WAF_LS'

    WB = openpyxl.Workbook()

    # 数据库审计巡检
    WAF_CHECKOUT(WB)

    WB.save(f'{data_path}/WAF_LS巡检信息.xlsx')

    print(f'输出的文件在：{data_path}')


"""
{
"code": 0,
"data": [
{
    "assignedtime": 31103036,
    "count": 1,
    "lincenseid": 174,
    "timestamp": "(2021-11-16 10:49:37)",
    "typeid": 2
},
{
    "assignedtime": 1728000000,
    "count": 1,
    "lincenseid": 171,
    "timestamp": "(2021-11-16 10:49:37)",
    "typeid": 1
}
]
}

"""
